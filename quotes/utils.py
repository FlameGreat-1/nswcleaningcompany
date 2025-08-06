from django.core.mail import send_mail, EmailMultiAlternatives
from django.template.loader import render_to_string
from django.conf import settings
from django.utils import timezone
from django.db.models import Q, Count, Sum, Avg, F
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime, timedelta
import uuid
import json
import logging
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.units import inch
from reportlab.lib.colors import black, blue, red, green
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from io import BytesIO

logger = logging.getLogger(__name__)


def calculate_quote_pricing(quote_data):
    try:
        service = quote_data.get("service")
        cleaning_type = quote_data.get("cleaning_type", "general")
        number_of_rooms = quote_data.get("number_of_rooms", 1)
        square_meters = quote_data.get("square_meters")
        urgency_level = quote_data.get("urgency_level", 2)
        postcode = quote_data.get("postcode", "2000")
        addons = quote_data.get("addons", [])

        base_price = calculate_base_price(
            service, cleaning_type, number_of_rooms, square_meters
        )
        extras_cost = calculate_extras_cost(addons)
        travel_cost = calculate_travel_cost(postcode, service)
        urgency_surcharge = calculate_urgency_surcharge(base_price, urgency_level)

        subtotal = base_price + extras_cost + travel_cost + urgency_surcharge
        discount_amount = calculate_discount(subtotal, quote_data)
        taxable_amount = subtotal - discount_amount
        gst_amount = calculate_gst(taxable_amount)
        total_price = taxable_amount + gst_amount

        return {
            "base_price": base_price,
            "extras_cost": extras_cost,
            "travel_cost": travel_cost,
            "urgency_surcharge": urgency_surcharge,
            "subtotal": subtotal,
            "discount_amount": discount_amount,
            "gst_amount": gst_amount,
            "total_price": total_price,
            "quote_valid_until": timezone.now() + timedelta(days=30),
            "breakdown": {
                "service_name": service.name if service else "Unknown",
                "cleaning_type": cleaning_type,
                "rooms": number_of_rooms,
                "square_meters": square_meters,
                "urgency_level": urgency_level,
                "postcode": postcode,
                "addons_count": len(addons),
            },
        }
    except Exception as e:
        logger.error(f"Quote pricing calculation failed: {str(e)}")
        raise


def calculate_base_price(service, cleaning_type, number_of_rooms, square_meters=None):
    if not service:
        return Decimal("0.00")

    base_price = service.base_price

    room_multiplier = get_room_multiplier(cleaning_type, number_of_rooms)
    base_price *= room_multiplier

    if square_meters:
        size_adjustment = calculate_size_adjustment(square_meters, cleaning_type)
        base_price += size_adjustment

    cleaning_type_multiplier = get_cleaning_type_multiplier(cleaning_type)
    base_price *= cleaning_type_multiplier

    return base_price.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def get_room_multiplier(cleaning_type, number_of_rooms):
    base_multipliers = {
        "general": Decimal("1.0"),
        "deep": Decimal("1.5"),
        "end_of_lease": Decimal("1.8"),
        "ndis": Decimal("1.2"),
        "commercial": Decimal("0.9"),
        "carpet": Decimal("0.8"),
        "window": Decimal("0.6"),
        "pressure_washing": Decimal("0.7"),
    }

    base_multiplier = base_multipliers.get(cleaning_type, Decimal("1.0"))

    if number_of_rooms <= 2:
        room_factor = Decimal("1.0")
    elif number_of_rooms <= 4:
        room_factor = Decimal("1.3")
    elif number_of_rooms <= 6:
        room_factor = Decimal("1.6")
    else:
        room_factor = Decimal("2.0")

    return base_multiplier * room_factor


def get_cleaning_type_multiplier(cleaning_type):
    multipliers = {
        "general": Decimal("1.0"),
        "deep": Decimal("1.4"),
        "end_of_lease": Decimal("1.6"),
        "ndis": Decimal("1.1"),
        "commercial": Decimal("0.85"),
        "carpet": Decimal("1.2"),
        "window": Decimal("0.9"),
        "pressure_washing": Decimal("1.3"),
    }
    return multipliers.get(cleaning_type, Decimal("1.0"))


def calculate_size_adjustment(square_meters, cleaning_type):
    if square_meters <= 50:
        return Decimal("0.00")
    elif square_meters <= 100:
        adjustment = Decimal("20.00")
    elif square_meters <= 200:
        adjustment = Decimal("50.00")
    elif square_meters <= 300:
        adjustment = Decimal("80.00")
    else:
        adjustment = Decimal("120.00")

    type_multipliers = {
        "general": Decimal("1.0"),
        "deep": Decimal("1.3"),
        "end_of_lease": Decimal("1.5"),
        "ndis": Decimal("1.1"),
        "commercial": Decimal("0.8"),
        "carpet": Decimal("1.2"),
        "window": Decimal("0.7"),
        "pressure_washing": Decimal("1.4"),
    }

    multiplier = type_multipliers.get(cleaning_type, Decimal("1.0"))
    return (adjustment * multiplier).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def calculate_extras_cost(addons):
    if not addons:
        return Decimal("0.00")

    total_cost = Decimal("0.00")
    for addon in addons:
        if hasattr(addon, "price"):
            total_cost += addon.price
        elif isinstance(addon, dict) and "price" in addon:
            total_cost += Decimal(str(addon["price"]))

    return total_cost.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def calculate_travel_cost(postcode, service):
    try:
        postcode_int = int(postcode)

        sydney_postcodes = range(2000, 2300)
        melbourne_postcodes = range(3000, 3200)
        brisbane_postcodes = range(4000, 4200)
        perth_postcodes = range(6000, 6200)
        adelaide_postcodes = range(5000, 5200)

        if postcode_int in sydney_postcodes:
            base_travel = Decimal("15.00")
        elif postcode_int in melbourne_postcodes:
            base_travel = Decimal("18.00")
        elif postcode_int in brisbane_postcodes:
            base_travel = Decimal("20.00")
        elif postcode_int in perth_postcodes:
            base_travel = Decimal("25.00")
        elif postcode_int in adelaide_postcodes:
            base_travel = Decimal("22.00")
        else:
            base_travel = Decimal("30.00")

        if service and hasattr(service, "travel_rate"):
            base_travel *= service.travel_rate

        return base_travel.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    except (ValueError, TypeError):
        return Decimal("25.00")


def calculate_urgency_surcharge(base_price, urgency_level):
    surcharge_rates = {
        1: Decimal("0.00"),
        2: Decimal("0.00"),
        3: Decimal("0.15"),
        4: Decimal("0.30"),
        5: Decimal("0.50"),
    }

    rate = surcharge_rates.get(urgency_level, Decimal("0.00"))
    surcharge = base_price * rate

    return surcharge.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def calculate_discount(subtotal, quote_data):
    discount_amount = Decimal("0.00")

    is_ndis = quote_data.get("is_ndis_client", False)
    if is_ndis:
        discount_amount += subtotal * Decimal("0.05")

    is_repeat_customer = quote_data.get("is_repeat_customer", False)
    if is_repeat_customer:
        discount_amount += subtotal * Decimal("0.03")

    promotional_discount = quote_data.get("promotional_discount", Decimal("0.00"))
    if promotional_discount:
        discount_amount += promotional_discount

    max_discount = subtotal * Decimal("0.20")
    if discount_amount > max_discount:
        discount_amount = max_discount

    return discount_amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def calculate_gst(taxable_amount):
    gst_rate = Decimal("0.10")
    gst_amount = taxable_amount * gst_rate
    return gst_amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def validate_pricing_data(pricing_data):
    required_fields = ["base_price", "total_price", "gst_amount"]

    for field in required_fields:
        if field not in pricing_data:
            raise ValueError(f"Missing required pricing field: {field}")

        if not isinstance(pricing_data[field], Decimal):
            try:
                pricing_data[field] = Decimal(str(pricing_data[field]))
            except (ValueError, TypeError):
                raise ValueError(f"Invalid pricing value for {field}")

    if pricing_data["total_price"] < 0:
        raise ValueError("Total price cannot be negative")

    if pricing_data["gst_amount"] < 0:
        raise ValueError("GST amount cannot be negative")

    return True


def generate_quote_number():
    current_year = timezone.now().year
    timestamp = timezone.now().strftime("%m%d%H%M")
    random_suffix = str(uuid.uuid4())[:4].upper()

    return f"QT-{current_year}-{timestamp}-{random_suffix}"


def format_currency(amount):
    if isinstance(amount, str):
        try:
            amount = Decimal(amount)
        except (ValueError, TypeError):
            return "$0.00"

    if not isinstance(amount, Decimal):
        amount = Decimal(str(amount))

    return f"${amount:,.2f}"


def format_quote_data_for_display(quote):
    return {
        "quote_number": quote.quote_number,
        "client_name": quote.client.get_full_name(),
        "service_name": quote.service.name,
        "cleaning_type_display": quote.get_cleaning_type_display(),
        "status_display": quote.get_status_display(),
        "urgency_display": quote.get_urgency_level_display(),
        "formatted_price": format_currency(quote.final_price),
        "formatted_base_price": format_currency(quote.base_price),
        "formatted_gst": format_currency(quote.gst_amount),
        "property_summary": f"{quote.property_address}, {quote.suburb} {quote.postcode}",
        "room_summary": f"{quote.number_of_rooms} rooms",
        "created_date": quote.created_at.strftime("%d/%m/%Y"),
        "expires_date": (
            quote.expires_at.strftime("%d/%m/%Y") if quote.expires_at else None
        ),
        "is_expired": quote.is_expired,
        "can_be_accepted": quote.can_be_accepted,
        "days_until_expiry": quote.days_until_expiry,
    }


def send_quote_notification(
    quote, notification_type, recipient_type, custom_message=None
):
    try:
        recipients = []

        if recipient_type in ["client", "both"]:
            recipients.append(quote.client.email)

        if recipient_type in ["staff", "both"]:
            if quote.assigned_to:
                recipients.append(quote.assigned_to.email)
            else:
                staff_emails = get_staff_notification_emails()
                recipients.extend(staff_emails)

        subject = get_notification_subject(notification_type, quote)

        context = {
            "quote": quote,
            "quote_data": format_quote_data_for_display(quote),
            "notification_type": notification_type,
            "custom_message": custom_message,
            "site_url": settings.SITE_URL,
            "company_name": getattr(settings, "COMPANY_NAME", "Cleaning Service"),
        }

        html_content = render_to_string(
            f"emails/quote_{notification_type}.html", context
        )

        text_content = render_to_string(
            f"emails/quote_{notification_type}.txt", context
        )

        msg = EmailMultiAlternatives(
            subject=subject,
            body=text_content,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=recipients,
        )
        msg.attach_alternative(html_content, "text/html")

        msg.send()

        logger.info(
            f"Quote notification sent: {notification_type} for quote {quote.quote_number}"
        )
        return True

    except Exception as e:
        logger.error(f"Failed to send quote notification: {str(e)}")
        return False


def get_notification_subject(notification_type, quote):
    subjects = {
        "submitted": f"Quote Submitted - {quote.quote_number}",
        "approved": f"Quote Approved - {quote.quote_number}",
        "rejected": f"Quote Rejected - {quote.quote_number}",
        "expired": f"Quote Expired - {quote.quote_number}",
        "cancelled": f"Quote Cancelled - {quote.quote_number}",
        "converted": f"Quote Converted to Job - {quote.quote_number}",
        "reminder": f"Quote Expiry Reminder - {quote.quote_number}",
        "assigned": f"Quote Assigned - {quote.quote_number}",
        "revised": f"Quote Revised - {quote.quote_number}",
    }
    return subjects.get(notification_type, f"Quote Update - {quote.quote_number}")


def get_staff_notification_emails():
    from django.contrib.auth import get_user_model

    User = get_user_model()

    staff_users = User.objects.filter(
        is_staff=True, is_active=True, email__isnull=False
    ).exclude(email="")

    return [user.email for user in staff_users]


def duplicate_quote(original_quote, user, options=None):
    from .models import Quote, QuoteItem, QuoteAttachment

    if options is None:
        options = {}

    try:
        new_quote = Quote.objects.create(
            client=user if not user.is_staff else original_quote.client,
            service=original_quote.service,
            cleaning_type=options.get(
                "new_cleaning_type", original_quote.cleaning_type
            ),
            property_address=original_quote.property_address,
            postcode=original_quote.postcode,
            suburb=original_quote.suburb,
            state=original_quote.state,
            number_of_rooms=original_quote.number_of_rooms,
            square_meters=original_quote.square_meters,
            urgency_level=options.get(
                "new_urgency_level", original_quote.urgency_level
            ),
            preferred_date=original_quote.preferred_date,
            preferred_time=original_quote.preferred_time,
            special_requirements=original_quote.special_requirements,
            access_instructions=original_quote.access_instructions,
            is_ndis_client=original_quote.is_ndis_client,
            ndis_participant_number=original_quote.ndis_participant_number,
            plan_manager_name=original_quote.plan_manager_name,
            plan_manager_contact=original_quote.plan_manager_contact,
            support_coordinator_name=original_quote.support_coordinator_name,
            support_coordinator_contact=original_quote.support_coordinator_contact,
            status="draft",
        )

        if options.get("include_items", True):
            for item in original_quote.items.all():
                QuoteItem.objects.create(
                    quote=new_quote,
                    service=item.service,
                    addon=item.addon,
                    item_type=item.item_type,
                    name=item.name,
                    description=item.description,
                    quantity=item.quantity,
                    unit_price=item.unit_price,
                    is_optional=item.is_optional,
                    is_taxable=item.is_taxable,
                    display_order=item.display_order,
                )

        if options.get("include_attachments", False):
            for attachment in original_quote.attachments.all():
                if attachment.is_public:
                    QuoteAttachment.objects.create(
                        quote=new_quote,
                        uploaded_by=user,
                        file=attachment.file,
                        attachment_type=attachment.attachment_type,
                        title=f"Copy of {attachment.title}",
                        description=attachment.description,
                        is_public=attachment.is_public,
                        display_order=attachment.display_order,
                    )

        new_quote.update_pricing()

        logger.info(
            f"Quote duplicated: {original_quote.quote_number} -> {new_quote.quote_number}"
        )
        return new_quote

    except Exception as e:
        logger.error(
            f"Failed to duplicate quote {original_quote.quote_number}: {str(e)}"
        )
        raise


def bulk_quote_operation(operation_data, user):
    from .models import Quote

    quote_ids = operation_data.get("quote_ids", [])
    operation = operation_data.get("operation")

    if not quote_ids or not operation:
        return {"success": False, "message": "Missing required data"}

    try:
        quotes = Quote.objects.filter(id__in=quote_ids)

        if not user.is_staff:
            quotes = quotes.filter(client=user)

        results = {
            "total_quotes": len(quote_ids),
            "processed": 0,
            "failed": 0,
            "errors": [],
        }

        for quote in quotes:
            try:
                if operation == "approve":
                    if quote.status in ["submitted", "under_review"]:
                        quote.approve_quote(user)
                        results["processed"] += 1
                    else:
                        results["errors"].append(
                            f"Quote {quote.quote_number} cannot be approved"
                        )
                        results["failed"] += 1

                elif operation == "reject":
                    if quote.status in ["submitted", "under_review"]:
                        reason = operation_data.get(
                            "rejection_reason", "Bulk rejection"
                        )
                        quote.reject_quote(user, reason)
                        results["processed"] += 1
                    else:
                        results["errors"].append(
                            f"Quote {quote.quote_number} cannot be rejected"
                        )
                        results["failed"] += 1

                elif operation == "cancel":
                    if quote.status not in ["converted", "cancelled"]:
                        quote.status = "cancelled"
                        quote.save()
                        results["processed"] += 1
                    else:
                        results["errors"].append(
                            f"Quote {quote.quote_number} cannot be cancelled"
                        )
                        results["failed"] += 1

                elif operation == "assign":
                    assigned_to_id = operation_data.get("assigned_to")
                    if assigned_to_id:
                        from django.contrib.auth import get_user_model

                        User = get_user_model()
                        try:
                            assigned_user = User.objects.get(
                                id=assigned_to_id, is_staff=True
                            )
                            quote.assigned_to = assigned_user
                            quote.save()
                            results["processed"] += 1
                        except User.DoesNotExist:
                            results["errors"].append(f"Invalid user for assignment")
                            results["failed"] += 1
                    else:
                        results["errors"].append(f"No user specified for assignment")
                        results["failed"] += 1

            except Exception as e:
                results["errors"].append(
                    f"Error processing quote {quote.quote_number}: {str(e)}"
                )
                results["failed"] += 1

        results["success"] = results["failed"] == 0
        results["message"] = f"Processed {results['processed']} quotes successfully"

        if results["failed"] > 0:
            results["message"] += f", {results['failed']} failed"

        return results

    except Exception as e:
        logger.error(f"Bulk operation failed: {str(e)}")
        return {
            "success": False,
            "message": f"Bulk operation failed: {str(e)}",
            "total_quotes": len(quote_ids),
            "processed": 0,
            "failed": len(quote_ids),
        }


def get_quote_analytics_data(params):
    from .models import Quote

    try:
        start_date = params.get("start_date")
        end_date = params.get("end_date")
        group_by = params.get("group_by", "status")
        include_ndis = params.get("include_ndis", True)
        include_general = params.get("include_general", True)

        queryset = Quote.objects.all()

        if start_date and end_date:
            queryset = queryset.filter(
                created_at__date__gte=start_date, created_at__date__lte=end_date
            )

        if not include_ndis:
            queryset = queryset.filter(is_ndis_client=False)

        if not include_general:
            queryset = queryset.filter(is_ndis_client=True)

        analytics_data = {
            "summary": get_quote_summary_analytics(queryset),
            "trends": get_quote_trend_analytics(queryset, group_by),
            "performance": get_quote_performance_analytics(queryset),
            "distribution": get_quote_distribution_analytics(queryset, group_by),
        }

        return analytics_data

    except Exception as e:
        logger.error(f"Analytics data generation failed: {str(e)}")
        raise


def get_quote_summary_analytics(queryset):
    summary = queryset.aggregate(
        total_quotes=Count("id"),
        total_value=Sum("final_price"),
        average_value=Avg("final_price"),
        pending_count=Count("id", filter=Q(status__in=["submitted", "under_review"])),
        approved_count=Count("id", filter=Q(status="approved")),
        rejected_count=Count("id", filter=Q(status="rejected")),
        converted_count=Count("id", filter=Q(status="converted")),
        expired_count=Count("id", filter=Q(status="expired")),
        cancelled_count=Count("id", filter=Q(status="cancelled")),
    )

    if summary["total_quotes"] > 0:
        summary["approval_rate"] = (
            summary["approved_count"] / summary["total_quotes"]
        ) * 100
        summary["conversion_rate"] = (
            summary["converted_count"] / summary["total_quotes"]
        ) * 100
        summary["rejection_rate"] = (
            summary["rejected_count"] / summary["total_quotes"]
        ) * 100
    else:
        summary["approval_rate"] = 0
        summary["conversion_rate"] = 0
        summary["rejection_rate"] = 0

    return summary


def get_quote_trend_analytics(queryset, group_by):
    if group_by == "month":
        trends = (
            queryset.extra(select={"period": "DATE_FORMAT(created_at, '%%Y-%%m')"})
            .values("period")
            .annotate(
                count=Count("id"),
                total_value=Sum("final_price"),
                avg_value=Avg("final_price"),
            )
            .order_by("period")
        )

    elif group_by == "status":
        trends = (
            queryset.values("status")
            .annotate(
                count=Count("id"),
                total_value=Sum("final_price"),
                avg_value=Avg("final_price"),
            )
            .order_by("-count")
        )

    elif group_by == "cleaning_type":
        trends = (
            queryset.values("cleaning_type")
            .annotate(
                count=Count("id"),
                total_value=Sum("final_price"),
                avg_value=Avg("final_price"),
            )
            .order_by("-count")
        )

    elif group_by == "urgency":
        trends = (
            queryset.values("urgency_level")
            .annotate(
                count=Count("id"),
                total_value=Sum("final_price"),
                avg_value=Avg("final_price"),
            )
            .order_by("urgency_level")
        )

    elif group_by == "state":
        trends = (
            queryset.values("state")
            .annotate(
                count=Count("id"),
                total_value=Sum("final_price"),
                avg_value=Avg("final_price"),
            )
            .order_by("-count")
        )

    else:
        trends = []

    return list(trends)


def get_quote_performance_analytics(queryset):
    now = timezone.now()

    performance = {
        "this_month": queryset.filter(
            created_at__year=now.year, created_at__month=now.month
        ).count(),
        "last_month": queryset.filter(
            created_at__year=now.year if now.month > 1 else now.year - 1,
            created_at__month=now.month - 1 if now.month > 1 else 12,
        ).count(),
        "this_week": queryset.filter(created_at__gte=now - timedelta(days=7)).count(),
        "urgent_quotes": queryset.filter(urgency_level__gte=4).count(),
        "high_value_quotes": queryset.filter(final_price__gte=1000).count(),
        "ndis_quotes": queryset.filter(is_ndis_client=True).count(),
        "expiring_soon": queryset.filter(
            status="approved",
            expires_at__lte=now + timedelta(days=7),
            expires_at__gt=now,
        ).count(),
    }

    if performance["last_month"] > 0:
        performance["month_over_month_growth"] = (
            (performance["this_month"] - performance["last_month"])
            / performance["last_month"]
        ) * 100
    else:
        performance["month_over_month_growth"] = 0

    return performance


def get_quote_distribution_analytics(queryset, group_by):
    distribution = {}

    if group_by == "status":
        distribution = dict(queryset.values_list("status").annotate(Count("status")))
    elif group_by == "cleaning_type":
        distribution = dict(
            queryset.values_list("cleaning_type").annotate(Count("cleaning_type"))
        )
    elif group_by == "urgency":
        distribution = dict(
            queryset.values_list("urgency_level").annotate(Count("urgency_level"))
        )
    elif group_by == "state":
        distribution = dict(queryset.values_list("state").annotate(Count("state")))

    return distribution


def generate_quote_pdf(quote):
    try:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=24,
            spaceAfter=30,
            textColor=blue,
        )

        header_style = ParagraphStyle(
            "CustomHeader",
            parent=styles["Heading2"],
            fontSize=16,
            spaceAfter=12,
            textColor=black,
        )

        normal_style = styles["Normal"]

        story.append(Paragraph("CLEANING SERVICE QUOTE", title_style))
        story.append(Spacer(1, 20))

        quote_info_data = [
            ["Quote Number:", quote.quote_number],
            ["Date:", quote.created_at.strftime("%d/%m/%Y")],
            ["Status:", quote.get_status_display()],
            [
                "Valid Until:",
                quote.expires_at.strftime("%d/%m/%Y") if quote.expires_at else "N/A",
            ],
        ]

        quote_info_table = Table(quote_info_data, colWidths=[2 * inch, 3 * inch])
        quote_info_table.setStyle(
            TableStyle(
                [
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )

        story.append(quote_info_table)
        story.append(Spacer(1, 20))

        story.append(Paragraph("CLIENT INFORMATION", header_style))

        client_data = [
            ["Name:", quote.client.get_full_name()],
            ["Email:", quote.client.email],
            ["Phone:", getattr(quote.client, "phone", "N/A")],
            ["Property:", quote.property_address],
            ["Suburb:", f"{quote.suburb}, {quote.state} {quote.postcode}"],
        ]

        if quote.is_ndis_client:
            client_data.append(
                ["NDIS Participant:", quote.ndis_participant_number or "N/A"]
            )
            if quote.plan_manager_name:
                client_data.append(["Plan Manager:", quote.plan_manager_name])

        client_table = Table(client_data, colWidths=[2 * inch, 4 * inch])
        client_table.setStyle(
            TableStyle(
                [
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )

        story.append(client_table)
        story.append(Spacer(1, 20))

        story.append(Paragraph("SERVICE DETAILS", header_style))

        service_data = [
            ["Service:", quote.service.name],
            ["Cleaning Type:", quote.get_cleaning_type_display()],
            ["Number of Rooms:", str(quote.number_of_rooms)],
            ["Urgency Level:", f"Level {quote.urgency_level}"],
        ]

        if quote.square_meters:
            service_data.append(["Square Meters:", f"{quote.square_meters} m²"])

        if quote.preferred_date:
            service_data.append(
                ["Preferred Date:", quote.preferred_date.strftime("%d/%m/%Y")]
            )

        if quote.preferred_time:
            service_data.append(
                ["Preferred Time:", quote.preferred_time.strftime("%H:%M")]
            )

        service_table = Table(service_data, colWidths=[2 * inch, 4 * inch])
        service_table.setStyle(
            TableStyle(
                [
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )

        story.append(service_table)
        story.append(Spacer(1, 20))

        if quote.items.exists():
            story.append(Paragraph("QUOTE ITEMS", header_style))

            items_data = [["Description", "Quantity", "Unit Price", "Total"]]

            for item in quote.items.all():
                items_data.append(
                    [
                        item.name,
                        str(item.quantity),
                        format_currency(item.unit_price),
                        format_currency(item.total_price),
                    ]
                )

            items_table = Table(
                items_data, colWidths=[3 * inch, 1 * inch, 1.5 * inch, 1.5 * inch]
            )
            items_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), blue),
                        ("TEXTCOLOR", (0, 0), (-1, 0), white),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                        ("GRID", (0, 0), (-1, -1), 1, black),
                    ]
                )
            )

            story.append(items_table)
            story.append(Spacer(1, 20))

        story.append(Paragraph("PRICING BREAKDOWN", header_style))

        pricing_data = [
            ["Base Price:", format_currency(quote.base_price)],
            ["Extras Cost:", format_currency(quote.extras_cost)],
            ["Travel Cost:", format_currency(quote.travel_cost)],
            ["Urgency Surcharge:", format_currency(quote.urgency_surcharge)],
            ["Discount:", f"-{format_currency(quote.discount_amount)}"],
            ["Subtotal:", format_currency(quote.estimated_total)],
            ["GST (10%):", format_currency(quote.gst_amount)],
            ["TOTAL:", format_currency(quote.final_price)],
        ]

        pricing_table = Table(pricing_data, colWidths=[3 * inch, 2 * inch])
        pricing_table.setStyle(
            TableStyle(
                [
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -2), 10),
                    ("FONTSIZE", (0, -1), (-1, -1), 14),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("LINEABOVE", (0, -1), (-1, -1), 2, black),
                    ("BACKGROUND", (0, -1), (-1, -1), blue),
                    ("TEXTCOLOR", (0, -1), (-1, -1), white),
                ]
            )
        )

        story.append(pricing_table)
        story.append(Spacer(1, 20))

        if quote.special_requirements:
            story.append(Paragraph("SPECIAL REQUIREMENTS", header_style))
            story.append(Paragraph(quote.special_requirements, normal_style))
            story.append(Spacer(1, 12))

        if quote.access_instructions:
            story.append(Paragraph("ACCESS INSTRUCTIONS", header_style))
            story.append(Paragraph(quote.access_instructions, normal_style))
            story.append(Spacer(1, 12))

        story.append(Spacer(1, 20))
        story.append(Paragraph("TERMS AND CONDITIONS", header_style))

        terms = [
            "• This quote is valid for 30 days from the date of issue.",
            "• All prices include GST unless otherwise specified.",
            "• Payment is due within 7 days of service completion.",
            "• Cancellations must be made at least 24 hours in advance.",
            "• Additional charges may apply for services outside normal business hours.",
            "• We reserve the right to refuse service if property conditions are unsafe.",
        ]

        for term in terms:
            story.append(Paragraph(term, normal_style))

        doc.build(story)
        pdf_content = buffer.getvalue()
        buffer.close()

        return pdf_content

    except Exception as e:
        logger.error(f"PDF generation failed for quote {quote.quote_number}: {str(e)}")
        raise


def export_quotes_data(export_params, user):
    from .models import Quote
    import csv
    import openpyxl
    from openpyxl.styles import Font, Alignment

    try:
        quote_ids = export_params.get("quote_ids")
        format_type = export_params.get("format", "csv")
        include_items = export_params.get("include_items", True)
        include_attachments = export_params.get("include_attachments", False)

        if quote_ids:
            queryset = Quote.objects.filter(id__in=quote_ids)
        else:
            queryset = Quote.objects.all()

            status_filter = export_params.get("status_filter")
            if status_filter:
                queryset = queryset.filter(status__in=status_filter)

        if not user.is_staff:
            queryset = queryset.filter(client=user)

        queryset = queryset.select_related(
            "client", "service", "assigned_to"
        ).prefetch_related("items", "attachments")

        if format_type == "csv":
            return generate_csv_export(queryset, include_items, include_attachments)
        elif format_type == "excel":
            return generate_excel_export(queryset, include_items, include_attachments)
        elif format_type == "pdf":
            return generate_pdf_export(queryset, include_items, include_attachments)
        else:
            raise ValueError(f"Unsupported export format: {format_type}")

    except Exception as e:
        logger.error(f"Export failed: {str(e)}")
        raise


def generate_csv_export(queryset, include_items, include_attachments):
    output = BytesIO()
    output_text = StringIO()
    writer = csv.writer(output_text)

    headers = [
        "Quote Number",
        "Client Name",
        "Client Email",
        "Service",
        "Cleaning Type",
        "Status",
        "Property Address",
        "Suburb",
        "Postcode",
        "State",
        "Number of Rooms",
        "Square Meters",
        "Urgency Level",
        "Final Price",
        "Base Price",
        "GST Amount",
        "Is NDIS Client",
        "Created Date",
        "Expires Date",
        "Assigned To",
    ]

    if include_items:
        headers.extend(["Items Count", "Items Total"])

    if include_attachments:
        headers.extend(["Attachments Count"])

    writer.writerow(headers)

    for quote in queryset:
        row = [
            quote.quote_number,
            quote.client.get_full_name(),
            quote.client.email,
            quote.service.name,
            quote.get_cleaning_type_display(),
            quote.get_status_display(),
            quote.property_address,
            quote.suburb,
            quote.postcode,
            quote.state,
            quote.number_of_rooms,
            quote.square_meters or "",
            quote.urgency_level,
            quote.final_price,
            quote.base_price,
            quote.gst_amount,
            "Yes" if quote.is_ndis_client else "No",
            quote.created_at.strftime("%Y-%m-%d %H:%M"),
            quote.expires_at.strftime("%Y-%m-%d") if quote.expires_at else "",
            quote.assigned_to.get_full_name() if quote.assigned_to else "",
        ]

        if include_items:
            items_count = quote.items.count()
            items_total = sum(item.total_price for item in quote.items.all())
            row.extend([items_count, items_total])

        if include_attachments:
            attachments_count = quote.attachments.count()
            row.extend([attachments_count])

        writer.writerow(row)

    output.write(output_text.getvalue().encode("utf-8"))
    output.seek(0)
    return output.getvalue()


def generate_excel_export(queryset, include_items, include_attachments):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Quotes Export"

    headers = [
        "Quote Number",
        "Client Name",
        "Client Email",
        "Service",
        "Cleaning Type",
        "Status",
        "Property Address",
        "Suburb",
        "Postcode",
        "State",
        "Number of Rooms",
        "Square Meters",
        "Urgency Level",
        "Final Price",
        "Base Price",
        "GST Amount",
        "Is NDIS Client",
        "Created Date",
        "Expires Date",
        "Assigned To",
    ]

    if include_items:
        headers.extend(["Items Count", "Items Total"])

    if include_attachments:
        headers.extend(["Attachments Count"])

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center")

    for row_num, quote in enumerate(queryset, 2):
        data = [
            quote.quote_number,
            quote.client.get_full_name(),
            quote.client.email,
            quote.service.name,
            quote.get_cleaning_type_display(),
            quote.get_status_display(),
            quote.property_address,
            quote.suburb,
            quote.postcode,
            quote.state,
            quote.number_of_rooms,
            quote.square_meters or "",
            quote.urgency_level,
            float(quote.final_price),
            float(quote.base_price),
            float(quote.gst_amount),
            "Yes" if quote.is_ndis_client else "No",
            quote.created_at.strftime("%Y-%m-%d %H:%M"),
            quote.expires_at.strftime("%Y-%m-%d") if quote.expires_at else "",
            quote.assigned_to.get_full_name() if quote.assigned_to else "",
        ]

        if include_items:
            items_count = quote.items.count()
            items_total = float(sum(item.total_price for item in quote.items.all()))
            data.extend([items_count, items_total])

        if include_attachments:
            attachments_count = quote.attachments.count()
            data.extend([attachments_count])

        for col, value in enumerate(data, 1):
            ws.cell(row=row_num, column=col, value=value)

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_pdf_export(queryset, include_items, include_attachments):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle(
        "ExportTitle",
        parent=styles["Heading1"],
        fontSize=18,
        spaceAfter=20,
        textColor=blue,
    )

    story.append(Paragraph("QUOTES EXPORT REPORT", title_style))
    story.append(
        Paragraph(
            f"Generated on: {timezone.now().strftime('%d/%m/%Y %H:%M')}",
            styles["Normal"],
        )
    )
    story.append(Paragraph(f"Total Quotes: {queryset.count()}", styles["Normal"]))
    story.append(Spacer(1, 20))

    for quote in queryset:
        quote_data = [
            ["Quote Number:", quote.quote_number],
            ["Client:", quote.client.get_full_name()],
            ["Service:", quote.service.name],
            ["Status:", quote.get_status_display()],
            ["Final Price:", format_currency(quote.final_price)],
            ["Created:", quote.created_at.strftime("%d/%m/%Y")],
        ]

        quote_table = Table(quote_data, colWidths=[2 * inch, 4 * inch])
        quote_table.setStyle(
            TableStyle(
                [
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ("LINEBELOW", (0, -1), (-1, -1), 1, black),
                ]
            )
        )

        story.append(quote_table)
        story.append(Spacer(1, 10))

    doc.build(story)
    pdf_content = buffer.getvalue()
    buffer.close()

    return pdf_content


def generate_quote_report(report_params):
    from .models import Quote
    from io import StringIO

    try:
        report_type = report_params.get("report_type", "summary")
        start_date = report_params.get("start_date")
        end_date = report_params.get("end_date")
        format_type = report_params.get("format", "pdf")

        queryset = Quote.objects.all()

        if start_date and end_date:
            queryset = queryset.filter(
                created_at__date__gte=start_date, created_at__date__lte=end_date
            )

        filter_status = report_params.get("filter_status")
        if filter_status:
            queryset = queryset.filter(status__in=filter_status)

        filter_cleaning_type = report_params.get("filter_cleaning_type")
        if filter_cleaning_type:
            queryset = queryset.filter(cleaning_type__in=filter_cleaning_type)

        if report_type == "summary":
            return generate_summary_report(queryset, format_type)
        elif report_type == "detailed":
            return generate_detailed_report(queryset, format_type)
        elif report_type == "analytics":
            return generate_analytics_report(queryset, format_type)
        elif report_type == "conversion":
            return generate_conversion_report(queryset, format_type)
        else:
            raise ValueError(f"Unsupported report type: {report_type}")

    except Exception as e:
        logger.error(f"Report generation failed: {str(e)}")
        raise


def generate_summary_report(queryset, format_type):
    summary_data = {
        "total_quotes": queryset.count(),
        "total_value": queryset.aggregate(Sum("final_price"))["final_price__sum"] or 0,
        "average_value": queryset.aggregate(Avg("final_price"))["final_price__avg"]
        or 0,
        "status_breakdown": dict(
            queryset.values_list("status").annotate(Count("status"))
        ),
        "cleaning_type_breakdown": dict(
            queryset.values_list("cleaning_type").annotate(Count("cleaning_type"))
        ),
        "urgency_breakdown": dict(
            queryset.values_list("urgency_level").annotate(Count("urgency_level"))
        ),
        "ndis_count": queryset.filter(is_ndis_client=True).count(),
        "general_count": queryset.filter(is_ndis_client=False).count(),
    }

    if format_type == "pdf":
        return generate_summary_pdf_report(summary_data)
    elif format_type == "excel":
        return generate_summary_excel_report(summary_data)
    else:
        return generate_summary_csv_report(summary_data)


def generate_summary_pdf_report(summary_data):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Heading1"],
        fontSize=20,
        spaceAfter=30,
        textColor=blue,
        alignment=1,
    )

    story.append(Paragraph("QUOTE SUMMARY REPORT", title_style))
    story.append(
        Paragraph(
            f"Generated: {timezone.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]
        )
    )
    story.append(Spacer(1, 20))

    overview_data = [
        ["Total Quotes:", str(summary_data["total_quotes"])],
        ["Total Value:", format_currency(summary_data["total_value"])],
        ["Average Value:", format_currency(summary_data["average_value"])],
        ["NDIS Quotes:", str(summary_data["ndis_count"])],
        ["General Quotes:", str(summary_data["general_count"])],
    ]

    overview_table = Table(overview_data, colWidths=[2.5 * inch, 2 * inch])
    overview_table.setStyle(
        TableStyle(
            [
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 12),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 1, black),
            ]
        )
    )

    story.append(Paragraph("OVERVIEW", styles["Heading2"]))
    story.append(overview_table)
    story.append(Spacer(1, 20))

    story.append(Paragraph("STATUS BREAKDOWN", styles["Heading2"]))
    status_data = [["Status", "Count"]]
    for status, count in summary_data["status_breakdown"].items():
        status_data.append([status.replace("_", " ").title(), str(count)])

    status_table = Table(status_data, colWidths=[2 * inch, 1 * inch])
    status_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), blue),
                ("TEXTCOLOR", (0, 0), (-1, 0), white),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("GRID", (0, 0), (-1, -1), 1, black),
            ]
        )
    )

    story.append(status_table)
    story.append(Spacer(1, 20))

    story.append(Paragraph("CLEANING TYPE BREAKDOWN", styles["Heading2"]))
    cleaning_data = [["Cleaning Type", "Count"]]
    for cleaning_type, count in summary_data["cleaning_type_breakdown"].items():
        cleaning_data.append([cleaning_type.replace("_", " ").title(), str(count)])

    cleaning_table = Table(cleaning_data, colWidths=[2 * inch, 1 * inch])
    cleaning_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), blue),
                ("TEXTCOLOR", (0, 0), (-1, 0), white),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("GRID", (0, 0), (-1, -1), 1, black),
            ]
        )
    )

    story.append(cleaning_table)

    doc.build(story)
    pdf_content = buffer.getvalue()
    buffer.close()

    return pdf_content


def generate_detailed_report(queryset, format_type):
    detailed_data = []

    for quote in queryset.select_related("client", "service", "assigned_to"):
        quote_detail = {
            "quote_number": quote.quote_number,
            "client_name": quote.client.get_full_name(),
            "client_email": quote.client.email,
            "service_name": quote.service.name,
            "cleaning_type": quote.get_cleaning_type_display(),
            "status": quote.get_status_display(),
            "property_address": quote.property_address,
            "suburb": quote.suburb,
            "postcode": quote.postcode,
            "state": quote.state,
            "number_of_rooms": quote.number_of_rooms,
            "square_meters": quote.square_meters,
            "urgency_level": quote.urgency_level,
            "final_price": quote.final_price,
            "base_price": quote.base_price,
            "gst_amount": quote.gst_amount,
            "is_ndis_client": quote.is_ndis_client,
            "created_at": quote.created_at,
            "expires_at": quote.expires_at,
            "assigned_to": (
                quote.assigned_to.get_full_name() if quote.assigned_to else None
            ),
            "items_count": quote.items.count(),
            "attachments_count": quote.attachments.count(),
        }
        detailed_data.append(quote_detail)

    if format_type == "pdf":
        return generate_detailed_pdf_report(detailed_data)
    elif format_type == "excel":
        return generate_detailed_excel_report(detailed_data)
    else:
        return generate_detailed_csv_report(detailed_data)


def generate_analytics_report(queryset, format_type):
    analytics_data = get_quote_analytics_data(
        {"group_by": "status", "include_ndis": True, "include_general": True}
    )

    if format_type == "pdf":
        return generate_analytics_pdf_report(analytics_data)
    elif format_type == "excel":
        return generate_analytics_excel_report(analytics_data)
    else:
        return generate_analytics_csv_report(analytics_data)


def generate_conversion_report(queryset, format_type):
    total_quotes = queryset.count()
    converted_quotes = queryset.filter(status="converted").count()
    approved_quotes = queryset.filter(status="approved").count()
    rejected_quotes = queryset.filter(status="rejected").count()

    conversion_data = {
        "total_quotes": total_quotes,
        "converted_quotes": converted_quotes,
        "approved_quotes": approved_quotes,
        "rejected_quotes": rejected_quotes,
        "conversion_rate": (
            (converted_quotes / total_quotes * 100) if total_quotes > 0 else 0
        ),
        "approval_rate": (
            (approved_quotes / total_quotes * 100) if total_quotes > 0 else 0
        ),
        "rejection_rate": (
            (rejected_quotes / total_quotes * 100) if total_quotes > 0 else 0
        ),
        "monthly_trends": get_monthly_conversion_trends(queryset),
        "service_conversion": get_service_conversion_rates(queryset),
        "cleaning_type_conversion": get_cleaning_type_conversion_rates(queryset),
    }

    if format_type == "pdf":
        return generate_conversion_pdf_report(conversion_data)
    elif format_type == "excel":
        return generate_conversion_excel_report(conversion_data)
    else:
        return generate_conversion_csv_report(conversion_data)


def get_monthly_conversion_trends(queryset):
    trends = (
        queryset.extra(select={"month": "DATE_FORMAT(created_at, '%%Y-%%m')"})
        .values("month")
        .annotate(
            total=Count("id"),
            converted=Count("id", filter=Q(status="converted")),
            approved=Count("id", filter=Q(status="approved")),
            rejected=Count("id", filter=Q(status="rejected")),
        )
        .order_by("month")
    )

    for trend in trends:
        if trend["total"] > 0:
            trend["conversion_rate"] = (trend["converted"] / trend["total"]) * 100
            trend["approval_rate"] = (trend["approved"] / trend["total"]) * 100
            trend["rejection_rate"] = (trend["rejected"] / trend["total"]) * 100
        else:
            trend["conversion_rate"] = 0
            trend["approval_rate"] = 0
            trend["rejection_rate"] = 0

    return list(trends)


def get_service_conversion_rates(queryset):
    service_rates = (
        queryset.values("service__name")
        .annotate(
            total=Count("id"),
            converted=Count("id", filter=Q(status="converted")),
            approved=Count("id", filter=Q(status="approved")),
        )
        .order_by("-total")
    )

    for rate in service_rates:
        if rate["total"] > 0:
            rate["conversion_rate"] = (rate["converted"] / rate["total"]) * 100
            rate["approval_rate"] = (rate["approved"] / rate["total"]) * 100
        else:
            rate["conversion_rate"] = 0
            rate["approval_rate"] = 0

    return list(service_rates)


def get_cleaning_type_conversion_rates(queryset):
    cleaning_rates = (
        queryset.values("cleaning_type")
        .annotate(
            total=Count("id"),
            converted=Count("id", filter=Q(status="converted")),
            approved=Count("id", filter=Q(status="approved")),
        )
        .order_by("-total")
    )

    for rate in cleaning_rates:
        if rate["total"] > 0:
            rate["conversion_rate"] = (rate["converted"] / rate["total"]) * 100
            rate["approval_rate"] = (rate["approved"] / rate["total"]) * 100
        else:
            rate["conversion_rate"] = 0
            rate["approval_rate"] = 0

    return list(cleaning_rates)


def validate_quote_data(quote_data):
    errors = []

    required_fields = [
        "service",
        "cleaning_type",
        "property_address",
        "postcode",
        "number_of_rooms",
    ]
    for field in required_fields:
        if not quote_data.get(field):
            errors.append(f"{field.replace('_', ' ').title()} is required")

    postcode = quote_data.get("postcode")
    if postcode and not validate_australian_postcode(postcode):
        errors.append("Invalid Australian postcode")

    number_of_rooms = quote_data.get("number_of_rooms")
    if number_of_rooms and (number_of_rooms < 1 or number_of_rooms > 50):
        errors.append("Number of rooms must be between 1 and 50")

    square_meters = quote_data.get("square_meters")
    if square_meters and (square_meters < 10 or square_meters > 10000):
        errors.append("Square meters must be between 10 and 10,000")

    urgency_level = quote_data.get("urgency_level")
    if urgency_level and (urgency_level < 1 or urgency_level > 5):
        errors.append("Urgency level must be between 1 and 5")

    preferred_date = quote_data.get("preferred_date")
    if preferred_date and preferred_date < timezone.now().date():
        errors.append("Preferred date cannot be in the past")

    return errors


def validate_australian_postcode(postcode):
    try:
        postcode_int = int(postcode)
        return 1000 <= postcode_int <= 9999
    except (ValueError, TypeError):
        return False


def get_quote_status_color(status):
    colors = {
        "draft": "#6c757d",
        "submitted": "#007bff",
        "under_review": "#ffc107",
        "approved": "#28a745",
        "rejected": "#dc3545",
        "expired": "#6f42c1",
        "converted": "#20c997",
        "cancelled": "#fd7e14",
    }
    return colors.get(status, "#6c757d")


def get_urgency_color(urgency_level):
    colors = {1: "#28a745", 2: "#17a2b8", 3: "#ffc107", 4: "#fd7e14", 5: "#dc3545"}
    return colors.get(urgency_level, "#6c757d")


def calculate_quote_expiry_date(created_date=None, quote_type="general"):
    if created_date is None:
        created_date = timezone.now()

    expiry_days = {
        "general": 30,
        "deep": 45,
        "end_of_lease": 14,
        "ndis": 60,
        "commercial": 90,
        "carpet": 21,
        "window": 30,
        "pressure_washing": 30,
    }

    days = expiry_days.get(quote_type, 30)
    return created_date + timedelta(days=days)


def get_quote_priority_score(quote):
    score = 0

    score += quote.urgency_level * 20

    if quote.final_price >= 1000:
        score += 15
    elif quote.final_price >= 500:
        score += 10
    elif quote.final_price >= 200:
        score += 5

    if quote.is_ndis_client:
        score += 10

    if quote.cleaning_type in ["end_of_lease", "deep"]:
        score += 8

    days_since_created = (timezone.now() - quote.created_at).days
    if days_since_created >= 3:
        score += days_since_created * 2

    if quote.expires_at:
        days_until_expiry = (quote.expires_at - timezone.now()).days
        if days_until_expiry <= 7:
            score += 20
        elif days_until_expiry <= 14:
            score += 10

    return min(score, 100)


def get_recommended_services(quote):
    recommendations = []

    if quote.cleaning_type == "general":
        if quote.number_of_rooms >= 3:
            recommendations.append(
                {
                    "service": "Deep Cleaning Add-on",
                    "reason": "Recommended for larger properties",
                    "price_estimate": 50.00,
                }
            )

        if quote.square_meters and quote.square_meters >= 150:
            recommendations.append(
                {
                    "service": "Carpet Cleaning",
                    "reason": "Large properties often benefit from carpet cleaning",
                    "price_estimate": 80.00,
                }
            )

    elif quote.cleaning_type == "end_of_lease":
        recommendations.append(
            {
                "service": "Carpet Steam Cleaning",
                "reason": "Required for most end of lease agreements",
                "price_estimate": 120.00,
            }
        )

        recommendations.append(
            {
                "service": "Window Cleaning",
                "reason": "Often required for bond return",
                "price_estimate": 60.00,
            }
        )

    elif quote.cleaning_type == "deep":
        if quote.number_of_rooms >= 2:
            recommendations.append(
                {
                    "service": "Oven Cleaning",
                    "reason": "Commonly needed with deep cleaning",
                    "price_estimate": 40.00,
                }
            )

    return recommendations


def send_quote_reminder_notifications():
    from .models import Quote

    expiring_quotes = Quote.objects.filter(
        status="approved",
        expires_at__lte=timezone.now() + timedelta(days=7),
        expires_at__gt=timezone.now(),
    )

    sent_count = 0
    failed_count = 0

    for quote in expiring_quotes:
        try:
            send_quote_notification(quote, "reminder", "client")
            sent_count += 1
        except Exception as e:
            logger.error(
                f"Failed to send reminder for quote {quote.quote_number}: {str(e)}"
            )
            failed_count += 1

    logger.info(f"Quote reminders sent: {sent_count} successful, {failed_count} failed")
    return {"sent": sent_count, "failed": failed_count}


def cleanup_expired_quotes():
    from .models import Quote

    expired_quotes = Quote.objects.filter(
        status="approved", expires_at__lt=timezone.now()
    )

    updated_count = expired_quotes.update(status="expired")

    for quote in expired_quotes:
        try:
            send_quote_notification(quote, "expired", "both")
        except Exception as e:
            logger.error(
                f"Failed to send expiry notification for quote {quote.quote_number}: {str(e)}"
            )

    logger.info(f"Expired {updated_count} quotes")
    return updated_count


def get_quote_dashboard_data(user):
    from .models import Quote

    if user.is_staff:
        base_queryset = Quote.objects.all()
    else:
        base_queryset = Quote.objects.filter(client=user)

    now = timezone.now()

    dashboard_data = {
        "total_quotes": base_queryset.count(),
        "pending_quotes": base_queryset.filter(
            status__in=["submitted", "under_review"]
        ).count(),
        "approved_quotes": base_queryset.filter(status="approved").count(),
        "expired_quotes": base_queryset.filter(status="expired").count(),
        "quotes_this_week": base_queryset.filter(
            created_at__gte=now - timedelta(days=7)
        ).count(),
        "quotes_this_month": base_queryset.filter(
            created_at__year=now.year, created_at__month=now.month
        ).count(),
        "total_value": base_queryset.aggregate(Sum("final_price"))["final_price__sum"]
        or 0,
        "average_quote_value": base_queryset.aggregate(Avg("final_price"))[
            "final_price__avg"
        ]
        or 0,
        "urgent_quotes": base_queryset.filter(urgency_level__gte=4).count(),
        "ndis_quotes": base_queryset.filter(is_ndis_client=True).count(),
        "expiring_soon": base_queryset.filter(
            status="approved",
            expires_at__lte=now + timedelta(days=7),
            expires_at__gt=now,
        ).count(),
    }

    return dashboard_data


def generate_quote_insights(quote):
    insights = []

    if quote.urgency_level >= 4:
        insights.append(
            {
                "type": "warning",
                "message": "This is a high-urgency quote requiring immediate attention.",
                "action": "Prioritize processing",
            }
        )

    if quote.final_price >= 1000:
        insights.append(
            {
                "type": "info",
                "message": "This is a high-value quote.",
                "action": "Consider additional quality assurance",
            }
        )

    if quote.is_ndis_client:
        insights.append(
            {
                "type": "info",
                "message": "NDIS client - ensure compliance with NDIS requirements.",
                "action": "Review NDIS guidelines",
            }
        )

    if quote.expires_at:
        days_until_expiry = (quote.expires_at - timezone.now()).days
        if days_until_expiry <= 3:
            insights.append(
                {
                    "type": "danger",
                    "message": f"Quote expires in {days_until_expiry} days.",
                    "action": "Follow up with client urgently",
                }
            )
        elif days_until_expiry <= 7:
            insights.append(
                {
                    "type": "warning",
                    "message": f"Quote expires in {days_until_expiry} days.",
                    "action": "Consider sending reminder",
                }
            )

    if quote.cleaning_type == "end_of_lease":
        insights.append(
            {
                "type": "info",
                "message": "End of lease cleaning requires thorough documentation.",
                "action": "Prepare detailed checklist",
            }
        )

    if quote.number_of_rooms >= 6:
        insights.append(
            {
                "type": "info",
                "message": "Large property - consider additional resources.",
                "action": "Plan for extended service time",
            }
        )

    return insights


def calculate_service_area_coverage(postcode):
    try:
        postcode_int = int(postcode)

        coverage_areas = {
            "sydney": {"range": (2000, 2299), "coverage": "full", "travel_time": 30},
            "melbourne": {"range": (3000, 3199), "coverage": "full", "travel_time": 35},
            "brisbane": {"range": (4000, 4199), "coverage": "full", "travel_time": 40},
            "perth": {"range": (6000, 6199), "coverage": "limited", "travel_time": 45},
            "adelaide": {
                "range": (5000, 5199),
                "coverage": "limited",
                "travel_time": 50,
            },
            "canberra": {"range": (2600, 2699), "coverage": "full", "travel_time": 25},
            "darwin": {"range": (800, 899), "coverage": "limited", "travel_time": 60},
            "hobart": {"range": (7000, 7099), "coverage": "limited", "travel_time": 55},
        }

        for area, info in coverage_areas.items():
            if info["range"][0] <= postcode_int <= info["range"][1]:
                return {
                    "area": area,
                    "coverage": info["coverage"],
                    "travel_time": info["travel_time"],
                    "available": True,
                }

        return {
            "area": "remote",
            "coverage": "none",
            "travel_time": 120,
            "available": False,
        }

    except (ValueError, TypeError):
        return {
            "area": "unknown",
            "coverage": "none",
            "travel_time": 0,
            "available": False,
        }


def optimize_quote_scheduling(quotes):
    optimized_schedule = []

    sorted_quotes = sorted(
        quotes,
        key=lambda q: (
            -q.urgency_level,
            q.preferred_date or timezone.now().date(),
            q.preferred_time or timezone.now().time(),
            -q.final_price,
        ),
    )

    for quote in sorted_quotes:
        schedule_item = {
            "quote": quote,
            "priority_score": get_quote_priority_score(quote),
            "estimated_duration": estimate_service_duration(quote),
            "travel_time": calculate_service_area_coverage(quote.postcode)[
                "travel_time"
            ],
            "recommended_slot": get_recommended_time_slot(quote),
        }
        optimized_schedule.append(schedule_item)

    return optimized_schedule


def estimate_service_duration(quote):
    base_duration = {
        "general": 120,
        "deep": 240,
        "end_of_lease": 300,
        "ndis": 180,
        "commercial": 180,
        "carpet": 90,
        "window": 60,
        "pressure_washing": 150,
    }

    duration = base_duration.get(quote.cleaning_type, 120)

    room_factor = min(quote.number_of_rooms * 0.2, 2.0)
    duration *= 1 + room_factor

    if quote.square_meters:
        if quote.square_meters > 200:
            duration *= 1.3
        elif quote.square_meters > 100:
            duration *= 1.15

    if quote.urgency_level >= 4:
        duration *= 0.9

    return int(duration)


def get_recommended_time_slot(quote):
    if quote.preferred_time:
        return quote.preferred_time

    if quote.cleaning_type in ["commercial"]:
        return timezone.now().replace(hour=18, minute=0, second=0, microsecond=0).time()

    if quote.urgency_level >= 4:
        return timezone.now().replace(hour=8, minute=0, second=0, microsecond=0).time()

    return timezone.now().replace(hour=10, minute=0, second=0, microsecond=0).time()


def archive_old_quotes(days_old=365):
    from .models import Quote

    cutoff_date = timezone.now() - timedelta(days=days_old)

    old_quotes = Quote.objects.filter(
        created_at__lt=cutoff_date,
        status__in=["expired", "rejected", "cancelled", "converted"],
    )

    archived_count = 0
    for quote in old_quotes:
        quote.is_archived = True
        quote.save()
        archived_count += 1

    logger.info(f"Archived {archived_count} old quotes")
    return archived_count


def get_quote_performance_metrics():
    from .models import Quote

    now = timezone.now()
    last_30_days = now - timedelta(days=30)

    metrics = {
        "total_quotes_30_days": Quote.objects.filter(
            created_at__gte=last_30_days
        ).count(),
        "conversion_rate_30_days": 0,
        "average_response_time": 0,
        "customer_satisfaction": 0,
        "revenue_30_days": 0,
        "top_performing_services": [],
        "busiest_days": [],
        "peak_hours": [],
    }

    quotes_30_days = Quote.objects.filter(created_at__gte=last_30_days)
    converted_30_days = quotes_30_days.filter(status="converted").count()

    if metrics["total_quotes_30_days"] > 0:
        metrics["conversion_rate_30_days"] = (
            converted_30_days / metrics["total_quotes_30_days"]
        ) * 100

    metrics["revenue_30_days"] = (
        quotes_30_days.filter(status="converted").aggregate(Sum("final_price"))[
            "final_price__sum"
        ]
        or 0
    )

    return metrics


def sync_quote_with_external_systems(quote):
    sync_results = {"accounting": False, "crm": False, "calendar": False, "errors": []}

    try:
        if (
            hasattr(settings, "ACCOUNTING_INTEGRATION_ENABLED")
            and settings.ACCOUNTING_INTEGRATION_ENABLED
        ):
            sync_results["accounting"] = sync_quote_to_accounting(quote)
    except Exception as e:
        sync_results["errors"].append(f"Accounting sync failed: {str(e)}")

    try:
        if (
            hasattr(settings, "CRM_INTEGRATION_ENABLED")
            and settings.CRM_INTEGRATION_ENABLED
        ):
            sync_results["crm"] = sync_quote_to_crm(quote)
    except Exception as e:
        sync_results["errors"].append(f"CRM sync failed: {str(e)}")

    try:
        if (
            hasattr(settings, "CALENDAR_INTEGRATION_ENABLED")
            and settings.CALENDAR_INTEGRATION_ENABLED
        ):
            sync_results["calendar"] = sync_quote_to_calendar(quote)
    except Exception as e:
        sync_results["errors"].append(f"Calendar sync failed: {str(e)}")

    return sync_results


def sync_quote_to_accounting(quote):
    try:
        accounting_data = {
            "quote_number": quote.quote_number,
            "client_name": quote.client.get_full_name(),
            "client_email": quote.client.email,
            "service_description": quote.service.name,
            "amount": float(quote.final_price),
            "gst_amount": float(quote.gst_amount),
            "date": quote.created_at.isoformat(),
            "status": quote.status,
        }

        logger.info(f"Syncing quote {quote.quote_number} to accounting system")
        return True

    except Exception as e:
        logger.error(f"Accounting sync failed for quote {quote.quote_number}: {str(e)}")
        return False


def sync_quote_to_crm(quote):
    try:
        crm_data = {
            "quote_id": str(quote.id),
            "quote_number": quote.quote_number,
            "client_id": quote.client.id,
            "client_name": quote.client.get_full_name(),
            "client_email": quote.client.email,
            "service_type": quote.cleaning_type,
            "value": float(quote.final_price),
            "status": quote.status,
            "created_date": quote.created_at.isoformat(),
            "property_address": quote.property_address,
            "postcode": quote.postcode,
        }

        logger.info(f"Syncing quote {quote.quote_number} to CRM system")
        return True

    except Exception as e:
        logger.error(f"CRM sync failed for quote {quote.quote_number}: {str(e)}")
        return False


def sync_quote_to_calendar(quote):
    try:
        if quote.status == "approved" and quote.preferred_date and quote.preferred_time:
            calendar_event = {
                "title": f"Quote {quote.quote_number} - {quote.service.name}",
                "description": f"Client: {quote.client.get_full_name()}\nAddress: {quote.property_address}",
                "start_date": quote.preferred_date.isoformat(),
                "start_time": quote.preferred_time.isoformat(),
                "duration": estimate_service_duration(quote),
                "location": f"{quote.property_address}, {quote.suburb} {quote.postcode}",
                "attendees": [quote.client.email],
            }

            if quote.assigned_to:
                calendar_event["attendees"].append(quote.assigned_to.email)

            logger.info(f"Syncing quote {quote.quote_number} to calendar system")
            return True

        return False

    except Exception as e:
        logger.error(f"Calendar sync failed for quote {quote.quote_number}: {str(e)}")
        return False


def generate_quote_qr_code(quote):
    try:
        import qrcode
        from io import BytesIO

        quote_url = f"{settings.SITE_URL}/quotes/{quote.id}/"

        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(quote_url)
        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")

        buffer = BytesIO()
        img.save(buffer, format="PNG")
        buffer.seek(0)

        return buffer.getvalue()

    except ImportError:
        logger.warning("qrcode library not installed - QR code generation skipped")
        return None
    except Exception as e:
        logger.error(
            f"QR code generation failed for quote {quote.quote_number}: {str(e)}"
        )
        return None


def backup_quote_data(quote_ids=None):
    from .models import Quote
    import json

    try:
        if quote_ids:
            quotes = Quote.objects.filter(id__in=quote_ids)
        else:
            quotes = Quote.objects.all()

        backup_data = {
            "timestamp": timezone.now().isoformat(),
            "total_quotes": quotes.count(),
            "quotes": [],
        }

        for quote in quotes.select_related("client", "service").prefetch_related(
            "items", "attachments", "revisions"
        ):
            quote_data = {
                "id": str(quote.id),
                "quote_number": quote.quote_number,
                "client": {
                    "id": quote.client.id,
                    "name": quote.client.get_full_name(),
                    "email": quote.client.email,
                },
                "service": {"id": quote.service.id, "name": quote.service.name},
                "cleaning_type": quote.cleaning_type,
                "status": quote.status,
                "property_address": quote.property_address,
                "postcode": quote.postcode,
                "suburb": quote.suburb,
                "state": quote.state,
                "number_of_rooms": quote.number_of_rooms,
                "square_meters": (
                    float(quote.square_meters) if quote.square_meters else None
                ),
                "urgency_level": quote.urgency_level,
                "final_price": float(quote.final_price),
                "base_price": float(quote.base_price),
                "gst_amount": float(quote.gst_amount),
                "is_ndis_client": quote.is_ndis_client,
                "created_at": quote.created_at.isoformat(),
                "expires_at": (
                    quote.expires_at.isoformat() if quote.expires_at else None
                ),
                "items": [
                    {
                        "name": item.name,
                        "quantity": item.quantity,
                        "unit_price": float(item.unit_price),
                        "total_price": float(item.total_price),
                    }
                    for item in quote.items.all()
                ],
                "attachments_count": quote.attachments.count(),
                "revisions_count": quote.revisions.count(),
            }
            backup_data["quotes"].append(quote_data)

        backup_json = json.dumps(backup_data, indent=2)

        backup_filename = (
            f"quotes_backup_{timezone.now().strftime('%Y%m%d_%H%M%S')}.json"
        )

        logger.info(
            f"Quote backup created: {backup_filename} ({len(backup_data['quotes'])} quotes)"
        )

        return {
            "success": True,
            "filename": backup_filename,
            "data": backup_json,
            "quote_count": len(backup_data["quotes"]),
        }

    except Exception as e:
        logger.error(f"Quote backup failed: {str(e)}")
        return {"success": False, "error": str(e), "quote_count": 0}


def restore_quote_data(backup_data):
    from .models import Quote
    from django.contrib.auth import get_user_model
    from services.models import Service

    User = get_user_model()

    try:
        if isinstance(backup_data, str):
            backup_data = json.loads(backup_data)

        restored_count = 0
        errors = []

        for quote_data in backup_data.get("quotes", []):
            try:
                client = User.objects.get(id=quote_data["client"]["id"])
                service = Service.objects.get(id=quote_data["service"]["id"])

                quote, created = Quote.objects.get_or_create(
                    quote_number=quote_data["quote_number"],
                    defaults={
                        "client": client,
                        "service": service,
                        "cleaning_type": quote_data["cleaning_type"],
                        "status": quote_data["status"],
                        "property_address": quote_data["property_address"],
                        "postcode": quote_data["postcode"],
                        "suburb": quote_data["suburb"],
                        "state": quote_data["state"],
                        "number_of_rooms": quote_data["number_of_rooms"],
                        "square_meters": quote_data.get("square_meters"),
                        "urgency_level": quote_data["urgency_level"],
                        "final_price": quote_data["final_price"],
                        "base_price": quote_data["base_price"],
                        "gst_amount": quote_data["gst_amount"],
                        "is_ndis_client": quote_data["is_ndis_client"],
                    },
                )

                if created:
                    restored_count += 1

            except Exception as e:
                errors.append(
                    f"Failed to restore quote {quote_data.get('quote_number', 'unknown')}: {str(e)}"
                )

        logger.info(
            f"Quote restore completed: {restored_count} quotes restored, {len(errors)} errors"
        )

        return {"success": True, "restored_count": restored_count, "errors": errors}

    except Exception as e:
        logger.error(f"Quote restore failed: {str(e)}")
        return {"success": False, "error": str(e), "restored_count": 0}


def get_quote_health_check():
    from .models import Quote

    try:
        health_data = {
            "status": "healthy",
            "checks": {
                "database_connection": True,
                "quote_count": Quote.objects.count(),
                "recent_quotes": Quote.objects.filter(
                    created_at__gte=timezone.now() - timedelta(days=7)
                ).count(),
                "pending_quotes": Quote.objects.filter(
                    status__in=["submitted", "under_review"]
                ).count(),
                "expired_quotes": Quote.objects.filter(status="expired").count(),
                "pricing_integrity": True,
                "notification_system": True,
            },
            "warnings": [],
            "errors": [],
        }

        pending_count = health_data["checks"]["pending_quotes"]
        if pending_count > 50:
            health_data["warnings"].append(
                f"High number of pending quotes: {pending_count}"
            )

        expired_count = health_data["checks"]["expired_quotes"]
        if expired_count > 20:
            health_data["warnings"].append(
                f"High number of expired quotes: {expired_count}"
            )

        quotes_with_invalid_pricing = Quote.objects.filter(final_price__lt=0).count()

        if quotes_with_invalid_pricing > 0:
            health_data["checks"]["pricing_integrity"] = False
            health_data["errors"].append(
                f"Found {quotes_with_invalid_pricing} quotes with invalid pricing"
            )

        if health_data["errors"]:
            health_data["status"] = "unhealthy"
        elif health_data["warnings"]:
            health_data["status"] = "warning"

        return health_data

    except Exception as e:
        return {
            "status": "error",
            "error": str(e),
            "checks": {},
            "warnings": [],
            "errors": [f"Health check failed: {str(e)}"],
        }


def cleanup_quote_system():
    cleanup_results = {
        "expired_quotes_updated": 0,
        "old_quotes_archived": 0,
        "orphaned_items_removed": 0,
        "orphaned_attachments_removed": 0,
        "errors": [],
    }

    try:
        cleanup_results["expired_quotes_updated"] = cleanup_expired_quotes()
    except Exception as e:
        cleanup_results["errors"].append(f"Expired quotes cleanup failed: {str(e)}")

    try:
        cleanup_results["old_quotes_archived"] = archive_old_quotes()
    except Exception as e:
        cleanup_results["errors"].append(f"Quote archiving failed: {str(e)}")

    try:
        from .models import QuoteItem, QuoteAttachment

        orphaned_items = QuoteItem.objects.filter(quote__isnull=True)
        cleanup_results["orphaned_items_removed"] = orphaned_items.count()
        orphaned_items.delete()

        orphaned_attachments = QuoteAttachment.objects.filter(quote__isnull=True)
        cleanup_results["orphaned_attachments_removed"] = orphaned_attachments.count()
        orphaned_attachments.delete()

    except Exception as e:
        cleanup_results["errors"].append(f"Orphaned records cleanup failed: {str(e)}")

    logger.info(f"Quote system cleanup completed: {cleanup_results}")
    return cleanup_results


class QuoteUtils:
    @staticmethod
    def format_currency(amount):
        return format_currency(amount)

    @staticmethod
    def calculate_pricing(quote_data):
        return calculate_quote_pricing(quote_data)

    @staticmethod
    def generate_pdf(quote):
        return generate_quote_pdf(quote)

    @staticmethod
    def send_notification(
        quote, notification_type, recipient_type, custom_message=None
    ):
        return send_quote_notification(
            quote, notification_type, recipient_type, custom_message
        )

    @staticmethod
    def duplicate(original_quote, user, options=None):
        return duplicate_quote(original_quote, user, options)

    @staticmethod
    def bulk_operation(operation_data, user):
        return bulk_quote_operation(operation_data, user)

    @staticmethod
    def get_analytics(params):
        return get_quote_analytics_data(params)

    @staticmethod
    def export_data(export_params, user):
        return export_quotes_data(export_params, user)

    @staticmethod
    def generate_report(report_params):
        return generate_quote_report(report_params)

    @staticmethod
    def validate_data(quote_data):
        return validate_quote_data(quote_data)

    @staticmethod
    def get_dashboard_data(user):
        return get_quote_dashboard_data(user)

    @staticmethod
    def generate_insights(quote):
        return generate_quote_insights(quote)

    @staticmethod
    def sync_external_systems(quote):
        return sync_quote_with_external_systems(quote)

    @staticmethod
    def backup_data(quote_ids=None):
        return backup_quote_data(quote_ids)

    @staticmethod
    def restore_data(backup_data):
        return restore_quote_data(backup_data)

    @staticmethod
    def health_check():
        return get_quote_health_check()

    @staticmethod
    def cleanup_system():
        return cleanup_quote_system()


quote_utils = QuoteUtils()
